#!/usr/bin/env python3
"""解析微信账单Excel文件"""

import openpyxl
from pathlib import Path
from datetime import datetime
import re


def parse_wechat_bill(filepath):
    """
    解析微信账单Excel文件
    
    返回格式：
    [
        {
            'date': '2026-07-24 17:56:21',
            'type': '亲属卡交易',
            'counterpart': '二娃',
            'product': '/',
            'income_expense': '支出',
            'amount': 2.5,
            'payment_method': '零钱',
            'status': '支付成功',
            'transaction_id': '4200003155202607248873035468',
            'merchant_id': '2784886972109187',
            'remark': '/'
        },
        ...
    ]
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    # 找到表头行（包含"交易时间"的行）
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        if row[0] and '交易时间' in str(row[0]):
            header_row = i
            break
    
    if not header_row:
        raise ValueError("未找到表头行")
    
    # 读取表头
    headers = []
    for cell in ws[header_row]:
        if cell.value:
            headers.append(str(cell.value).strip())
    
    # 读取数据行
    records = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        # 跳过空行
        if not row or not row[0]:
            continue
        
        # 转换为字符串并清理
        values = [str(cell).strip() if cell is not None else '' for cell in row]
        
        # 确保列数匹配
        if len(values) < len(headers):
            values.extend([''] * (len(headers) - len(values)))
        
        # 构建记录
        record = {}
        for i, header in enumerate(headers):
            record[header] = values[i] if i < len(values) else ''
        
        # 转换金额
        amount_str = record.get('金额(元)', '0')
        # 移除货币符号
        amount_str = re.sub(r'[¥￥\s]', '', amount_str)
        try:
            record['金额(元)'] = float(amount_str)
        except:
            record['金额(元)'] = 0.0
        
        records.append(record)
    
    return records


def filter_records(records, filter_type=None, start_date=None, end_date=None):
    """
    过滤记录
    
    filter_type: '收入' / '支出' / None(全部)
    start_date: '2026-01-01' 格式
    end_date: '2026-12-31' 格式
    """
    filtered = records
    
    if filter_type:
        filtered = [r for r in filtered if r.get('收/支') == filter_type]
    
    if start_date:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        filtered = [r for r in filtered if datetime.strptime(r['交易时间'].split()[0], '%Y-%m-%d') >= start_dt]
    
    if end_date:
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
        filtered = [r for r in filtered if datetime.strptime(r['交易时间'].split()[0], '%Y-%m-%d') <= end_dt]
    
    return filtered


def summary(records):
    """生成统计摘要"""
    income = sum(r['金额(元)'] for r in records if r.get('收/支') == '收入')
    expense = sum(r['金额(元)'] for r in records if r.get('收/支') == '支出')
    
    return {
        'total_records': len(records),
        'income_count': len([r for r in records if r.get('收/支') == '收入']),
        'income_total': income,
        'expense_count': len([r for r in records if r.get('收/支') == '支出']),
        'expense_total': expense,
        'net': income - expense
    }


def parse_alipay_bill(filepath):
    """
    解析支付宝账单CSV文件
    
    支付宝CSV格式（GBK编码）：
    交易时间,交易分类,交易对方,商品名称,金额,收/支,交易状态,交易订单号,商家订单号,备注
    
    返回格式与微信账单一致
    """
    import csv
    
    # 支付宝CSV使用GBK编码
    with open(filepath, 'r', encoding='gbk', errors='ignore') as f:
        content = f.read()
    
    # 移除可能的BOM和汇总行
    lines = content.strip().split('\n')
    
    # 找到表头行（包含"交易时间"的行）
    header_row_idx = None
    for i, line in enumerate(lines):
        if '交易时间' in line:
            header_row_idx = i
            break
    
    if header_row_idx is None:
        raise ValueError("未找到表头行")
    
    # 解析CSV
    reader = csv.DictReader(lines[header_row_idx:])
    
    records = []
    for row in reader:
        # 跳过空行或汇总行
        if not row.get('交易时间') or '---' in row.get('交易时间', ''):
            continue
        
        # 转换金额
        amount_str = row.get('金额', '0')
        amount_str = re.sub(r'[¥￥\s]', '', amount_str)
        try:
            amount = float(amount_str)
        except:
            amount = 0.0
        
        # 映射支付宝字段到统一格式
        record = {
            '交易时间': row.get('交易时间', ''),
            '交易类型': row.get('交易分类', ''),
            '交易对方': row.get('交易对方', ''),
            '商品': row.get('商品名称', ''),
            '收/支': row.get('收/支', ''),
            '金额(元)': amount,
            '支付方式': '支付宝',  # 支付宝账单默认
            '当前状态': row.get('交易状态', ''),
            '交易单号': row.get('交易订单号', '').strip(),
            '商家单号': row.get('商家订单号', '').strip(),
            '备注': row.get('备注', '')
        }
        
        records.append(record)
    
    return records


if __name__ == "__main__":
    # 测试解析
    xlsx_file = 'data/wechat_bill_extracted/微信支付账单流水文件(20260424-20260724)——【解压密码可在微信支付公众号查看】.xlsx'
    
    print(f"解析: {xlsx_file}\n")
    records = parse_wechat_bill(xlsx_file)
    
    print(f"共解析 {len(records)} 条记录\n")
    
    # 统计
    stats = summary(records)
    print("=" * 60)
    print("统计摘要：")
    print(f"  总记录数: {stats['total_records']}")
    print(f"  收入: {stats['income_count']}笔, {stats['income_total']:.2f}元")
    print(f"  支出: {stats['expense_count']}笔, {stats['expense_total']:.2f}元")
    print(f"  净额: {stats['net']:.2f}元")
    print("=" * 60)
    
    # 显示前5条
    print("\n前5条记录：")
    for i, record in enumerate(records[:5], 1):
        print(f"\n{i}. {record['交易时间']}")
        print(f"   类型: {record['交易类型']}")
        print(f"   对方: {record['交易对方']}")
        print(f"   商品: {record['商品']}")
        print(f"   收/支: {record['收/支']}")
        print(f"   金额: {record['金额(元)']:.2f}元")
        print(f"   支付方式: {record['支付方式']}")
        print(f"   状态: {record['当前状态']}")
